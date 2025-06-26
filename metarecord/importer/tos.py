import logging
import re
from collections import OrderedDict

from openpyxl import load_workbook

from metarecord.models import (
    Action,
    Attribute,
    AttributeValue,
    Classification,
    Function,
    Phase,
    Record,
)


class TOSImporterException(Exception):  # noqa: N818
    def __init__(self, message, row_num=None):
        self.message = message
        self.row_num = row_num

    def __str__(self):
        message = self.message
        if self.row_num is not None:
            message = "{}: {}".format(self.row_num, self.message)
        return message


class TOSImporter:
    # valid attributes and their identifiers
    CHOICE_ATTRIBUTES = {
        "Julkisuusluokka": "PublicityClass",
        "Salassapitoaika": "SecurityPeriod",
        "Salassapitoperuste": "SecurityReason",
        "Henkilötietoluonne": "PersonalData",
        "Säilytysaika": "RetentionPeriod",
        "Säilytysajan peruste": "RetentionReason",
        "Suojeluluokka": "ProtectionClass",
        "Henkilötunnus": "SocialSecurityNumber",
        "Säilytysajan laskentaperuste": "RetentionPeriodStart",
        "Paperiasiakirjojen säilytysjärjestys": "StorageOrder",
        "Paperiasiakirjojen säilytysaika arkistossa": "RetentionPeriodTotal",
        "Paperiasiakirjojen säilytysaika työpisteessä": "RetentionPeriodOffice",
        "Salassapitoajan laskentaperuste": "Restriction.SecurityPeriodStart",
        "Suojaustaso": "Restriction.ProtectionLevel",
        "Turvallisuusluokka": "Restriction.SecurityClass",
        "Asiakirjatyyppi": "RecordType",
    }
    FREE_TEXT_ATTRIBUTES = {
        "Lisätietoja": "AdditionalInformation",
        "Rekisteröinti/ Tietojärjestelmä": "InformationSystem",
        "Paperiasiakirjojen säilytyspaikka": "StorageLocation",
        "Paperiasiakirjojen säilytyksen vastuuhenkilö": "StorageAccountable",
    }

    ALL_ATTRIBUTES = dict(CHOICE_ATTRIBUTES, **FREE_TEXT_ATTRIBUTES)

    MODEL_MAPPING = OrderedDict(
        [
            ("Asian metatiedot", Function),
            ("Käsittelyvaiheen metatiedot", Phase),
            ("Toimenpiteen metatiedot", Action),
            ("Asiakirjan metatiedot", Record),
            ("Asiakirjan liitteen metatiedot", Record),
        ]
    )
    MODEL_HIERARCHY = list(MODEL_MAPPING.values())

    ATTACHMENT_RECORD_TYPE_NAME = "liite"

    def __init__(self, options=None):
        self.options = options or {}
        self.wb = None
        self.logger = logging.getLogger(__name__)

    def open(self, excel):
        self.wb = load_workbook(excel, read_only=True)

    def _emit_warning(self, text, row_num=None):
        if row_num is not None:
            text = "{}: {}".format(row_num, text)
        self.logger.warning(text)

    def _clean_header(self, s):
        if not s:
            return s
        # strip leading numeric ids
        s = re.sub(r"^[0-9. ]+", "", s)
        # strip stuff within parentheses
        s = re.sub(r" \(.+\)", "", s)
        # convert multiple spaces into one
        s = re.sub(r" {2,}", " ", s)
        # if there's a '=' in the header, remove the last part
        s = s.split("=")[0].strip()
        return s

    def _clean_attribute_value(self, s):
        if s is None:
            return None
        return re.sub(r"\s\s+", " ", str(s)).strip() or None

    def _get_codesets(self, sheet):
        header_row = 5
        value_row = 5 + 1

        max_col = sheet.max_column + 1
        attr_names = [
            sheet.cell(row=header_row, column=x).value for x in range(1, max_col)
        ]

        codesets = {}

        for col, attr in enumerate(attr_names):
            attr = self._clean_header(attr)

            # fix inconsistencies in headers between codesets and the actual data
            if attr == "Paperiasiakirjojen säilytysaika työpisteeessä":
                attr = "Paperiasiakirjojen säilytysaika työpisteessä"
            elif attr == "Rekisteröinti/tietojärjestelmä":
                attr = "Rekisteröinti/ Tietojärjestelmä"

            codesets[attr] = []
            for row in range(value_row, sheet.max_row + 1):
                val = sheet.cell(row=row, column=col + 1).value
                if val is None:
                    break

                # strip stuff within parentheses from value
                val = re.sub(r" \(.+\)", "", str(val))

                codesets[attr].append(val)

        return codesets

    def _get_first_data_row(self, sheet):
        data_row = 1
        while sheet.cell(row=data_row, column=1).value != "Asian metatiedot":
            data_row += 1
            if data_row > sheet.max_row:
                raise TOSImporterException("Cannot find first data row")
        return data_row

    def _get_data(self, sheet):
        data_row = self._get_first_data_row(sheet)
        if not data_row:
            return None

        max_col = sheet.max_column + 1
        headers = [sheet.cell(row=1, column=x).value for x in range(1, max_col)]
        headers = [self._clean_header(s) for s in headers if s]
        data = []
        cells = sheet.iter_rows(
            min_row=data_row,
            max_row=sheet.max_row + 1,
            min_col=1,
            max_col=max_col,
        )
        for row in cells:
            attrs = {}
            for col, attr in enumerate(headers):
                cleaned_value = self._clean_attribute_value(row[col].value)
                if cleaned_value is not None:
                    attrs[attr] = cleaned_value
            data.append(attrs)
        return data

    def _get_classification_code(self, sheet):
        first_data_row = self._get_first_data_row(sheet)
        if not first_data_row:
            return None

        # get all four function id hierarchy levels
        # for example ['5', '05 01', '05 01 03', None]
        classification_codes = []
        for row in range(2, first_data_row):
            value = sheet.cell(row=row, column=1).value
            if value:
                value = str(value).strip()
            classification_codes.append(value)

        # find the last index before none (or the last one in the list if none doesn't
        # exist) that is the index of this function's id
        index = len(classification_codes) - 1
        while classification_codes[index] is None and index > 0:
            index -= 1

        return str(classification_codes[index])

    def _get_classification(self, code):
        queryset = Classification.objects.filter(code=code)

        if not queryset.exists():
            raise TOSImporterException("Classification %s does not exist" % code)

        classification = queryset.latest_approved().first()
        if not classification:
            classification = queryset.latest_version().first()
        return classification

    def _import_function(self, sheet):
        classification_code = self._get_classification_code(sheet)
        if not classification_code:
            return

        classification = self._get_classification(classification_code)
        Function.objects.filter(classification=classification).delete()

        if not classification.function_allowed:
            print(  # noqa: T201
                "Skipping, classification %s does not allow function creation."
                % classification_code
            )
            return

        return Function.objects.create(classification=classification)

    def _clean_attributes(self, original_attribute_data, row_num=None):
        cleaned_attribute_data = {}

        for attribute_name, attribute_value in original_attribute_data.items():
            if attribute_name == "Asiakirjan tyyppi":
                attribute_name = "Asiakirjatyyppi"
            if attribute_name in (
                "Säilytysaika",
                "Paperiasiakirjojen säilytysaika arkistossa",
            ) and str(attribute_value).startswith("-1"):
                attribute_value = "-1"
            if attribute_name not in self.ALL_ATTRIBUTES:
                self._emit_warning('Illegal attribute: "%s"' % attribute_name, row_num)
                continue
            if attribute_name == "Paperiasiakirjojen säilytysaika arkistossa":
                continue
            if (attribute_name, attribute_value) in (
                ("Suojeluluokka", "Ei suojeluluokkaa, sähköinen asiakirja"),
                ("Paperiasiakirjojen säilytysjärjestys", "Säilytetään sähköisesti"),
                (
                    "Paperiasiakirjojen säilytysaika työpisteessä",
                    "Säilytetään sähköisesti",
                ),
            ):
                continue

            cleaned_attribute_data[self.ALL_ATTRIBUTES[attribute_name]] = str(
                attribute_value
            )

        return cleaned_attribute_data

    def _save_structural_element(self, model, parent, data, index, parent_record=None):
        model_attributes = {}  # model specific attributes

        # attribute and value describing name. most commonly the attribute is
        # TypeSpecifier, but it might be PhaseType for Phase or ActionType for action
        name_attribute = {"TypeSpecifier": data["name"]}

        if model == Phase:
            parent_field_name = "function"
            # if the value is a valid PhaseType attribute value then PhaseType
            # attribute is used for name,
            # otherwise TypeSpecifier
            is_phase_type = AttributeValue.objects.filter(
                attribute__identifier="PhaseType", value=data["name"]
            ).exists()
            if is_phase_type:
                name_attribute = {"PhaseType": data["name"]}
        elif model == Action:
            parent_field_name = "phase"
            # if the value is a valid ActionType attribute value then ActionType
            # attribute is used for name,
            # otherwise TypeSpecifier
            is_action_type = AttributeValue.objects.filter(
                attribute__identifier="ActionType", value=data["name"]
            ).exists()
            if is_action_type:
                name_attribute = {"ActionType": data["name"]}
        elif model == Record and parent_record is None:
            parent_field_name = "action"
        else:  # attachment
            parent_field_name = "action"
            model_attributes["parent"] = parent_record

        model_attributes[parent_field_name] = parent

        model_attributes["index"] = index

        new_obj = model(**model_attributes)

        new_obj.attributes = name_attribute
        new_obj.attributes.update(data["attributes"])

        new_obj.save()
        return new_obj

    def _save_function(self, function):
        function_obj = function["obj"]

        # Make sure children are nuked
        function_obj.phases.all().delete()

        for idx, phase in enumerate(function["phases"], 1):
            phase_obj = self._save_structural_element(Phase, function_obj, phase, idx)
            for idx, action in enumerate(phase["actions"], 1):
                action_obj = self._save_structural_element(
                    Action, phase_obj, action, idx
                )
                record_idx = 0
                for record in action["records"]:
                    record_idx += 1
                    record_obj = self._save_structural_element(
                        Record, action_obj, record, record_idx
                    )
                    for attachment in record["attachments"]:
                        record_idx += 1
                        self._save_structural_element(
                            Record, action_obj, attachment, record_idx, record_obj
                        )

        function_obj.attributes = function["attributes"]
        function_obj.error_count = function.get("error_count", 0)
        function_obj.save()

    def _process_data(self, sheet, function_obj):  # noqa: C901
        data = self._get_data(sheet)

        if not data:
            return

        first_data_row = self._get_first_data_row(sheet)

        # Parse data with a state machine
        function = {"phases": []}
        target = function
        self.current_function = function
        phase = action = None
        previous = None

        for row_num, row in enumerate(data, first_data_row):
            name = None
            child_list = []

            if not row:
                continue

            try:
                type_info = str(row.pop("Tehtäväluokka")).strip()
            except KeyError:
                raise TOSImporterException("Cannot determine target", row_num)

            if type_info == "Asian metatiedot":
                target["obj"] = function_obj
                name = function_obj.name
                # Must be the first row
                assert phase is None and action is None
                previous = Function
            elif type_info.startswith("Asiakirjallisen tiedon käsittely"):
                assert phase is None and action is None
                # Skip row
                continue
            elif type_info == "Käsittelyvaiheen metatiedot":
                if previous is None:
                    raise TOSImporterException(
                        "Parent of phase %s missing" % row.get("Käsittelyvaihe"),
                        row_num,
                    )
                phase = {}
                child_list = function["phases"]
                phase["actions"] = []
                action = None
                target = phase
                name = row.pop("Käsittelyvaihe", None)
                assert name
                previous = Phase
            elif type_info == "Toimenpiteen metatiedot":
                if previous == Function:
                    raise TOSImporterException(
                        "Parent of action %s missing" % row.get("Toimenpide"), row_num
                    )
                action = {}
                child_list = phase["actions"]
                action["records"] = []
                target = action
                name = row.pop("Toimenpide", None)
                previous = Action
            elif type_info == "Asiakirjan metatiedot":
                if previous in (Function, Phase):
                    raise TOSImporterException(
                        "Parent of record %s missing"
                        % row.get("Asiakirjatyypin tarkenne"),
                        row_num,
                    )
                record = {}
                child_list = action["records"]
                record["attachments"] = []
                target = record
                name = row.pop("Asiakirjatyypin tarkenne", None)
                previous = Record
            elif type_info == "Asiakirjan liitteen metatiedot":
                if previous in (Function, Phase, Action):
                    raise TOSImporterException(
                        "Parent of attachment %s missing"
                        % row.get("Asiakirjan liitteet"),
                        row_num,
                    )
                attachment = {}
                child_list = record["attachments"]
                target = attachment
                name = row.pop("Asiakirjan liitteet", "---")
                row.pop("Asiakirjatyypin tarkenne", None)

            if type_info not in self.MODEL_MAPPING:
                raise TOSImporterException("Unknown type %s" % type_info, row_num)
            target_model = self.MODEL_MAPPING[type_info]

            if target_model != Function and (not name or len(name) <= 2):
                if row:
                    raise TOSImporterException("Cannot determine name", row_num)
                continue

            target["name"] = name

            # Clean some attributes
            for name in ("Säilytysaika", "Paperiasiakirjojen säilytysaika arkistossa"):
                s = row.get(name)
                if isinstance(s, str) and s.startswith("-1"):
                    row[name] = "-1"

            target["attributes"] = self._clean_attributes(row, row_num)
            child_list.append(target)
        self._save_function(function)

    def import_attributes(self):  # noqa: C901
        self.logger.info("Importing attributes...")

        try:
            sheet = self.wb["Koodistot"]
        except KeyError:
            self.logger.info(
                "Cannot import attributes, the workbook does not contain sheet"
                ' "Koodistot".'
            )
            return

        codesets = self._get_codesets(sheet)
        handled_attrs = set()

        for attr, values in codesets.items():
            self.logger.info("\nProcessing %s" % attr)

            if attr == "Asiakirjatyypit":
                attr = "Asiakirjatyyppi"

            if attr not in self.ALL_ATTRIBUTES:
                self.logger.info("    skipping ")
                continue

            try:
                attribute_obj, created = Attribute.objects.update_or_create(
                    identifier=self.ALL_ATTRIBUTES.get(attr), defaults={"name": attr}
                )
            except ValueError as e:
                self.logger.info("    !!!! Cannot create attribute: %s" % e)
                continue
            handled_attrs.add(self.ALL_ATTRIBUTES.get(attr))

            if attr in self.FREE_TEXT_ATTRIBUTES:
                self.logger.info("    free text attribute")
                continue

            for value in values:
                try:
                    cleaned_value = self._clean_attribute_value(value)
                    if not cleaned_value:
                        raise ValueError('Invalid value: "%s"' % value)
                    obj, created = AttributeValue.objects.get_or_create(
                        attribute=attribute_obj, value=cleaned_value
                    )
                except ValueError as e:
                    self.logger.info("    !!!! Cannot create attribute value: %s" % e)
                    continue

                info_str = "Created" if created else "Already exist"
                self.logger.info("    %s: %s" % (info_str, value))

        # add also free text attributes that don't exist in the codesets sheet
        for name, identifier in self.FREE_TEXT_ATTRIBUTES.items():
            if identifier not in handled_attrs:
                attribute_obj, created = Attribute.objects.update_or_create(
                    identifier=identifier, defaults={"name": name}
                )
                info_str = "Created" if created else "Already exist"
                self.logger.info(
                    "\n%s: free text attribute %s that doesn't exist in the sheet"
                    % (info_str, name)
                )

        self.logger.info("\nDone.")

    def import_data(self):
        self.logger.info("Importing data...")
        for sheet in self.wb:
            try:
                self.logger.info("Processing sheet %s" % sheet.title)
                if (
                    sheet.max_column <= 2
                    or sheet.max_row <= 2
                    or sheet["A1"].value != "Tehtäväluokka"
                    or "ahjo" in str(sheet["A2"].value).lower()
                ):
                    self.logger.info("Skipping")
                    continue

                # process function
                function = self._import_function(sheet)

                # process data
                if function:
                    self._process_data(sheet, function)
            except TOSImporterException as e:
                if self.options.get("ignore_errors"):
                    self.logger.info("Skipping, got exception: %s" % e)
                else:
                    raise

        self.logger.info("Done.")

    def import_template(self, sheet_name, template_name):
        self.logger.info("Importing template...")

        sheet = self.wb.get_sheet_by_name(sheet_name)
        template_name = template_name or sheet_name

        function, created = Function.objects.get_or_create(
            name=template_name, is_template=True
        )

        if created:
            self.logger.info("Creating new template %s" % template_name)
        else:
            self.logger.info("Updating template %s" % template_name)

        self._process_data(sheet, function)

        self.logger.info("Done.")
